from openpyxl import load_workbook

ROW_ID_COLUMN = "_row_id"


def values_differ(original_value, uploaded_value) -> bool:
    """
    Compare two cell values for a real difference, tolerant of numeric formatting
    differences (e.g. 45000000.0 vs 45000000) that aren't actual edits made by the auditor.
    """
    orig_str = "" if original_value is None else str(original_value).strip()
    up_str = "" if uploaded_value is None else str(uploaded_value).strip()

    if orig_str == up_str:
        return False

    # Try numeric comparison so formatting differences aren't treated as real edits
    try:
        if float(orig_str) == float(up_str):
            return False
    except (ValueError, TypeError):
        pass

    return True


def diff_uploaded_against_snapshot(uploaded_file_path: str, snapshot_rows: list) -> dict:
    """
    Read a re-uploaded, auditor-edited Excel file (from the cleaning workbook export) and
    compare it against the saved snapshot of what was originally downloaded.

    Rows are matched by the hidden "_row_id" column written into every export, NOT by
    position. This means the auditor can freely delete rows (e.g. a row they judged to be
    unreliable or unnecessary) and the system still correctly identifies exactly which
    original row is missing, rather than misaligning the comparison or rejecting the file
    outright just because the row count changed.

    Columns are compared by name. A column present in the snapshot but absent from the
    uploaded file is reported as a deleted column — this is surfaced to the caller as
    information, not silently blocked, since deciding to drop an unreliable column is a
    legitimate audit judgment call. The caller (the endpoint) decides what to do with that
    information, e.g. recording it as part of the audit trail.

    Returns a dict with:
      "corrections": list of {row_index, column, original_value, corrected_value} for every
        data cell that actually changed.
      "deleted_row_ids": list of original row indices present in the snapshot but missing
        from the uploaded file — rows the auditor removed.
      "deleted_columns": list of column names present in the snapshot but missing from the
        uploaded file — columns the auditor removed entirely.
      "header_renames": dict of {new_column_name: True} for any column whose header was
        still marked [UNRESOLVED] in the export and has since been renamed by the auditor.
      "new_row_ids": list of row ids found in the uploaded file that were never in the
        snapshot — this should not normally happen since the hidden id column is locked
        from being repurposed, but is reported for safety.
    """
    wb = load_workbook(uploaded_file_path, data_only=True)
    if "Cleaned Data" not in wb.sheetnames:
        raise ValueError("Uploaded file does not contain a 'Cleaned Data' sheet. Please upload the file exactly as it was downloaded.")
    ws = wb["Cleaned Data"]

    raw_headers = [cell.value for cell in ws[1]]
    if not raw_headers or raw_headers[0] != ROW_ID_COLUMN:
        raise ValueError(
            "Uploaded file is missing its row tracking column. Please upload the file "
            "exactly as downloaded, without removing or modifying the hidden first column."
        )

    # Strip the [UNRESOLVED] prefix to get each column's real underlying name, while
    # tracking which positions were unresolved so renames can be detected. Skip the
    # row-id column (position 0) and Issues column from this column list.
    clean_headers = []
    was_unresolved = []
    for h in raw_headers[1:]:
        if h == "Issues":
            continue
        if h and str(h).startswith("[UNRESOLVED]"):
            clean_headers.append(str(h).replace("[UNRESOLVED]", "", 1).strip())
            was_unresolved.append(True)
        else:
            clean_headers.append(h)
            was_unresolved.append(False)

    # Build a lookup of uploaded rows keyed by their row id, reading every data row
    uploaded_rows_by_id = {}
    for row_cells in ws.iter_rows(min_row=2, values_only=False):
        row_id_value = row_cells[0].value
        if row_id_value is None:
            continue  # skip any fully blank trailing row
        row_id = int(row_id_value)
        row_data = {}
        for col_offset, col_name in enumerate(clean_headers, start=1):
            row_data[col_name] = row_cells[col_offset].value
        uploaded_rows_by_id[row_id] = row_data

    # Close the workbook now that every value needed has been read out of it. On Windows,
    # openpyxl keeps the underlying file handle open until close() is called explicitly —
    # without this, the caller's attempt to delete the temp upload file right after this
    # function returns fails with PermissionError: file in use by another process.
    wb.close()

    # Snapshot columns (excluding the internal _row_index key)
    snapshot_columns = set()
    snapshot_rows_by_id = {}
    for snap_row in snapshot_rows:
        row_id = snap_row["_row_index"]
        snapshot_rows_by_id[row_id] = snap_row
        snapshot_columns.update(k for k in snap_row.keys() if k != "_row_index")

    uploaded_columns = set(clean_headers)

    # Columns present in the snapshot but missing from the upload entirely = deleted columns.
    # This is reported, not blocked — removing an unreliable column is a legitimate audit choice.
    deleted_columns = sorted(snapshot_columns - uploaded_columns)

    # Columns present in the upload but never in the snapshot — likely a renamed
    # previously-unresolved column. Reported separately from a genuine new/unexpected column.
    new_columns = sorted(uploaded_columns - snapshot_columns)
    header_renames = {col: True for col in new_columns}

    # Rows present in the snapshot but missing from the upload entirely = deliberately deleted rows
    deleted_row_ids = sorted(set(snapshot_rows_by_id.keys()) - set(uploaded_rows_by_id.keys()))

    # Rows present in the upload that were never in the snapshot — should not normally
    # happen since the id column is locked, reported for safety/debugging
    new_row_ids = sorted(set(uploaded_rows_by_id.keys()) - set(snapshot_rows_by_id.keys()))

    # Diff cell values only for rows that exist in both, only for columns that exist in both
    shared_columns = snapshot_columns & uploaded_columns
    corrections = []
    for row_id, snap_row in snapshot_rows_by_id.items():
        if row_id not in uploaded_rows_by_id:
            continue  # this row was deleted, already captured in deleted_row_ids
        uploaded_row = uploaded_rows_by_id[row_id]
        for col_name in shared_columns:
            original_value = snap_row.get(col_name)
            uploaded_value = uploaded_row.get(col_name)
            if values_differ(original_value, uploaded_value):
                corrections.append({
                    "row_index": row_id,
                    "column": col_name,
                    "original_value": "" if original_value is None else str(original_value),
                    "corrected_value": "" if uploaded_value is None else str(uploaded_value),
                })

    return {
        "corrections": corrections,
        "deleted_row_ids": deleted_row_ids,
        "deleted_columns": deleted_columns,
        "header_renames": header_renames,
        "new_row_ids": new_row_ids,
    }