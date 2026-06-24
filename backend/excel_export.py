from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Fill colors used consistently across the workbook.
# Only ONE fill color marks "this row has an issue" — no severity-based color variation.
FLAGGED_ROW_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # red highlight
HEADER_FILL = PatternFill(start_color="FF1E3A5F", end_color="FF1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
UNRESOLVED_HEADER_FONT = Font(color="FFFFFFFF", bold=True, italic=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")

# Name of the hidden column that stores each row's original dataframe index. This lets a
# later re-uploaded file be matched back to the correct row by stable ID rather than by
# position, so the auditor can delete or reorder rows in Excel and the system still knows
# exactly which original row each remaining row corresponds to.
ROW_ID_COLUMN = "_row_id"


def _is_unresolved_column(mapping: dict, original_col: str) -> bool:
    """
    A column is unresolved if its mapping entry still says mapped_to == "unknown".
    These columns were never renamed by clean_dataframe and still sit under their
    original name in the cleaned data — we mark them visibly so the auditor can
    finish that mapping decision while looking at the real data in Excel.
    """
    info = mapping.get(original_col)
    return isinstance(info, dict) and info.get("mapped_to") == "unknown"


def _group_issues_by_row(issues: list) -> dict:
    """
    Group row-level issues by their row_index so we can tell which rows are flagged.
    Column-level issues (row_index == "N/A") are excluded — they don't belong to a
    specific row and only appear on the Issues Summary sheet.
    """
    grouped = {}
    for issue in issues:
        row_index = issue.get("row_index")
        if row_index == "N/A" or row_index is None:
            continue
        grouped.setdefault(int(row_index), []).append(issue)
    return grouped


def build_cleaning_workbook(cleaned_df, report: dict, mapping: dict, filename_hint: str = "cleaned_data") -> BytesIO:
    """
    Build a two-sheet workbook for the cleaning correction loop:
      Sheet 1 "Cleaned Data" — the cleaned dataframe exactly as-is. Rows with at least
        one issue are highlighted in red so they're easy to spot. Columns still mapped
        to "unknown" get their header marked "[UNRESOLVED] <original name>" so the
        auditor can finish that mapping decision while looking at real data. No Issues
        column here — full issue details live on the Issues Summary sheet instead, so
        this sheet stays focused purely on the data itself.
        A hidden "_row_id" column stores each row's original dataframe index, so a
        later re-uploaded file can be matched back to the correct row even if rows
        were deleted or reordered by the auditor in Excel.
        The sheet is NOT protected/locked — the auditor can freely edit any cell,
        including headers, rows, and columns, since audit work requires that flexibility.
      Sheet 2 "Issues Summary" — a flat checklist of every issue (including
        column-level issues like unknown columns and sparse-column flags), with the
        row number so the auditor can cross-reference back to Sheet 1.
    Returns an in-memory BytesIO of the .xlsx file, ready to stream as a download.
    """
    issues = report.get("issues", [])
    issues_by_row = _group_issues_by_row(issues)

    wb = Workbook()

    # ---------- Sheet 1: Cleaned Data ----------
    data_sheet = wb.active
    data_sheet.title = "Cleaned Data"

    columns = list(cleaned_df.columns)
    # Drop the internal duplicate-marker column if present — it's not meant for the auditor to see
    columns = [c for c in columns if c != "_is_duplicate"]
    # Hidden row-id column goes first, then all data columns. No Issues column —
    # that detail lives on the Issues Summary sheet only.
    header_row = [ROW_ID_COLUMN] + columns

    # Write header row. Unresolved columns get a visibly different label and font style
    # so they stand out from columns that are already confirmed.
    for col_idx, col_name in enumerate(header_row, start=1):
        cell = data_sheet.cell(row=1, column=col_idx)
        if col_name != ROW_ID_COLUMN and _is_unresolved_column(mapping, col_name):
            cell.value = f"[UNRESOLVED] {col_name}"
            cell.font = UNRESOLVED_HEADER_FONT
        else:
            cell.value = col_name
            cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = TOP_ALIGNMENT

    # Write data rows. dataframe index is used directly as the row id, matching the
    # issue's row_index produced by clean_dataframe since both come from the same dataframe.
    for row_offset, (df_index, row) in enumerate(cleaned_df.iterrows(), start=2):
        row_issues = issues_by_row.get(df_index, [])
        is_flagged = len(row_issues) > 0

        # Hidden row id, written first in every row
        id_cell = data_sheet.cell(row=row_offset, column=1)
        id_cell.value = int(df_index)
        id_cell.alignment = TOP_ALIGNMENT
        if is_flagged:
            id_cell.fill = FLAGGED_ROW_FILL

        for col_offset, col_name in enumerate(columns, start=2):
            cell = data_sheet.cell(row=row_offset, column=col_offset)
            value = row[col_name]
            cell.value = "" if (value is None or str(value).lower() == "nan") else value
            cell.alignment = TOP_ALIGNMENT
            if is_flagged:
                cell.fill = FLAGGED_ROW_FILL

    # Column widths — set explicitly based on content so headers and data line up consistently.
    # The hidden row-id column doesn't need a visible width since it's hidden anyway.
    data_sheet.column_dimensions[get_column_letter(1)].width = 8
    data_sheet.column_dimensions[get_column_letter(1)].hidden = True
    for col_offset, col_name in enumerate(columns, start=2):
        header_len = len(f"[UNRESOLVED] {col_name}") if _is_unresolved_column(mapping, col_name) else len(col_name)
        max_data_len = max([len(str(v)) for v in cleaned_df[col_name].head(50)] or [0])
        width = min(max(header_len, max_data_len) + 4, 40)
        data_sheet.column_dimensions[get_column_letter(col_offset)].width = width

    # Freeze the header row so it stays visible while scrolling through data
    data_sheet.freeze_panes = "A2"

    # No sheet protection — the auditor needs to freely edit cells, rows, and columns,
    # including occasionally renaming an [UNRESOLVED] header. Protection caused real
    # editing problems in Excel and isn't worth the tradeoff for an audit tool where
    # flexibility matters more than guarding against accidental renames.

    # ---------- Sheet 2: Issues Summary ----------
    summary_sheet = wb.create_sheet("Issues Summary")
    summary_headers = ["Row", "Column", "Issue", "Severity"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for row_offset, issue in enumerate(issues, start=2):
        summary_sheet.cell(row=row_offset, column=1).value = issue.get("row", "N/A")
        summary_sheet.cell(row=row_offset, column=2).value = issue.get("column", "")
        issue_cell = summary_sheet.cell(row=row_offset, column=3)
        issue_cell.value = issue.get("issue", "")
        issue_cell.alignment = WRAP_ALIGNMENT
        summary_sheet.cell(row=row_offset, column=4).value = issue.get("severity", "")

    summary_sheet.column_dimensions["A"].width = 10
    summary_sheet.column_dimensions["B"].width = 28
    summary_sheet.column_dimensions["C"].width = 70
    summary_sheet.column_dimensions["D"].width = 12
    summary_sheet.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

    